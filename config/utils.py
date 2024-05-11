## NOTE:- For Download Excel File
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from django.http import FileResponse
from io import BytesIO

class ExcelDataDownload:
    def __init__(self, excel_data, filename):
        self.excel_data = excel_data
        self.filename = filename

    def generate_response(self):
        workbook = Workbook()
        sheet = workbook.active

        # Set the background color for the header row
        header_fill = PatternFill(start_color="3f86ea", end_color="3f86ea", fill_type="solid")

        sheet.append(self.excel_data[0])  # Append the header row to the sheet

        # Apply the background color to the header row
        for cell in sheet[1]:
            cell.fill = header_fill

        # Append the rest of the data
        for row in self.excel_data[1:]:
            sheet.append(row)

        # Save the workbook to a BytesIO buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        # Create a FileResponse with the buffer
        response = FileResponse(
            buffer,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={self.filename}.xlsx'

        return response



## NOTE:- For Pagination
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import Http404

class CustomPaginator:
    def __init__(self, queryset, obj_per_page):
        self.paginator = Paginator(queryset, obj_per_page)

    def get_paginated_data(self, request):
        page_number = request.GET.get('page', 1)

        try:
            page_obj = self.paginator.get_page(page_number)
        except EmptyPage:
            raise Http404("Page not found")

        num_pages = self.paginator.num_pages
        current_page = page_obj.number

        start_page = max(1, min(current_page - 2, num_pages - 4))
        end_page = min(num_pages + 1, start_page + 5)
        page_range = range(start_page, end_page)

        return {
            'page_obj': page_obj,
            'page_range': page_range,
        }
    


def percentage(part, whole):
    Percentage = 100 * float(part)/float(whole)
    # return str(Percentage) + "%"
    return Percentage


from django.forms.models import model_to_dict
from django.db.models.fields.files import ImageFieldFile


def model_to_dict_with_nulls(instance):
    opts = instance._meta
    data = {}
    for f in opts.fields:
        if hasattr(instance, f.name):
            value = getattr(instance, f.name)
            data[f.name] = value
    return data

# def model_to_dict_with_nulls(instance):
#     opts = instance._meta
#     data = {}
#     for f in opts.fields:
#         if hasattr(instance, f.name):
#             value = getattr(instance, f.name)
#             if isinstance(value, ImageFieldFile):
#                 # Convert ImageFieldFile to its URL
#                 value = value.url if value else None
#             data[f.name] = value
#     return data
